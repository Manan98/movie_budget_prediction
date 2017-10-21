import json,requests,csv
from openpyxl import load_workbook

wb2 = load_workbook('movie_list.xlsx')
ws = wb2['movie_list']
i=1
ws.cell(row=i, column=1).value="budget"
ws.cell(row=i, column=2).value="revenue"
ws.cell(row=i, column=3).value="profit"
ws.cell(row=i, column=4).value="popularity"
ws.cell(row=i, column=5).value="runtime"
ws.cell(row=i, column=6).value="vote_average"
ws.cell(row=i, column=7).value="genre"
with open('movie_list.xlsx','r') as f:
    writer=csv.writer(f,dialect='excel')
    for num in range(50000):
        url="https://api.themoviedb.org//3/movie/"+str(num)+"?language=en-US&api_key=27304d0cacd43611ff1d8050d749367f"
        r=requests.get(url)
        if(r.status_code==200):
            r=r.json()
            #print(type(str(r)))
            dat=json.dumps(r)
            data=json.loads(dat)
            if(data["budget"]!=0 and data["revenue"]!=0):
                i+=1
                ws.cell(row=i, column=1).value = data["budget"]
                ws.cell(row=i, column=2).value = data["revenue"]
                ws.cell(row=i, column=3).value = str(((int)(data["revenue"])-(int)(data["budget"])))
                ws.cell(row=i, column=4).value = data["popularity"]
                ws.cell(row=i, column=5).value = data["runtime"]
                ws.cell(row=i, column=6).value = data["vote_average"]
                try:
                    ws.cell(row=i, column=7).value = data["genres"][0]["id"]
                except:
                    pass
wb2.save('movie_list.xlsx')